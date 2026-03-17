"""
Export all frontrunning detection results to a multi-sheet Excel workbook.
Sheets: Summary, Sandwich Attacks, Victims, Displacement, Arbitrage, Suppression
"""
import os
from datetime import datetime

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import PAIR_CONFIG

# ── styles ──────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
PROFIT_POS  = Font(name='Arial', color='27AE60', size=10)
PROFIT_NEG  = Font(name='Arial', color='E74C3C', size=10)
NORMAL_FONT = Font(name='Arial', size=10)
LINK_FONT   = Font(name='Arial', size=10, color='2980B9', underline='single')
TITLE_FONT  = Font(name='Arial', bold=True, size=14)
SUB_FONT    = Font(name='Arial', bold=True, size=11, color='2C3E50')
THIN_BORDER = Border(bottom=Side(style='thin', color='D5D8DC'))
NUM_FMT_USD  = '#,##0.00'
NUM_FMT_GWEI = '#,##0.00'


def _auto_width(ws, min_w=10, max_w=46):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        lengths = [len(str(c.value)) for c in col_cells if c.value is not None]
        ws.column_dimensions[letter].width = max(min_w, min(max_w, max(lengths, default=min_w) + 2))


def _write_header(ws, headers, row=1):
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = f'A{row + 1}'


def _write_rows(ws, df, headers, row_builder, start_row=2):
    """Generic helper: for each DataFrame row call row_builder(row) → list of values."""
    for ri, (_, data_row) in enumerate(df.iterrows(), start=start_row):
        vals = row_builder(data_row)
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER


def _tx_url(h):
    return f'https://etherscan.io/tx/{h}'


def _human(raw, decimals):
    if pd.isna(raw) or raw == 0:
        return 0.0
    return round(raw / (10 ** decimals), 6)


def _ts(t):
    try:
        return datetime.utcfromtimestamp(int(t)).strftime('%Y-%m-%d %H:%M:%S')
    except Exception:
        return ''


# ══════════════════════════════════════════════════════════════════
#  Public API
# ══════════════════════════════════════════════════════════════════

def export_excel(results: dict, output_dir: str) -> str:
    wb = Workbook()
    _build_summary(wb, results)
    _build_sandwich_sheet(wb, results.get('sandwiches', pd.DataFrame()))
    _build_victim_sheet(wb, results.get('victims', pd.DataFrame()))
    _build_displacement_sheet(wb, results.get('displacement', pd.DataFrame()))
    _build_arbitrage_sheet(wb, results.get('arbitrage', pd.DataFrame()))
    _build_suppression_sheet(wb, results.get('suppression', pd.DataFrame()))
    path = os.path.join(output_dir, 'frontrunning_analysis.xlsx')
    wb.save(path)
    return path


# ── Summary ─────────────────────────────────────────────────────

def _build_summary(wb, res):
    ws = wb.active
    ws.title = 'Summary'
    ws.cell(1, 1, 'MEV Frontrunning Analysis — Summary').font = TITLE_FONT
    ws.cell(2, 1, f'Generated: {datetime.now():%Y-%m-%d %H:%M:%S}').font = NORMAL_FONT

    sw  = res.get('sandwiches', pd.DataFrame())
    dp  = res.get('displacement', pd.DataFrame())
    arb = res.get('arbitrage', pd.DataFrame())
    sup = res.get('suppression', pd.DataFrame())
    vic = res.get('victims', pd.DataFrame())

    r = 4
    stats = [
        ('Frontrunning Type', 'Count'),
        ('Sandwich (Insertion)', f'{len(sw):,}'),
        ('Displacement', f'{len(dp):,}'),
        ('Arbitrage / Back-run', f'{len(arb):,}'),
        ('Suppression', f'{len(sup):,}'),
        ('', ''),
        ('Total Victim Transactions', f'{len(vic):,}'),
        ('Unique Sandwich Attackers', f'{sw["attacker"].nunique():,}' if not sw.empty else '0'),
    ]

    if not sw.empty and 'net_profit_usd' in sw.columns:
        stats += [
            ('', ''),
            ('Sandwich Gross Profit (USD)', f'${sw["profit_usd"].sum():,.2f}'),
            ('Sandwich Net Profit (USD)',   f'${sw["net_profit_usd"].sum():,.2f}'),
            ('Avg Sandwich Net Profit',     f'${sw["net_profit_usd"].mean():,.2f}'),
            ('Max Sandwich Net Profit',     f'${sw["net_profit_usd"].max():,.2f}'),
        ]

    if not arb.empty and 'net_profit_usd' in arb.columns:
        stats += [
            ('', ''),
            ('Arbitrage Net Profit (USD)', f'${arb["net_profit_usd"].sum():,.2f}'),
        ]

    for label, value in stats:
        ws.cell(r, 1, label).font = SUB_FONT if label else NORMAL_FONT
        ws.cell(r, 2, value).font = NORMAL_FONT
        r += 1

    # Per-pair table
    r += 1
    ws.cell(r, 1, 'Breakdown by Pair').font = TITLE_FONT
    r += 1
    pair_h = ['Pair', 'Sandwiches', 'Displacement', 'Arbitrage', 'Suppression']
    _write_header(ws, pair_h, row=r)
    r += 1
    pairs = sorted(set(
        list(sw['pair'].unique() if not sw.empty else [])
        + list(dp['pair'].unique() if not dp.empty else [])
        + list(arb['pair'].unique() if not arb.empty else [])
        + list(sup['pair'].unique() if not sup.empty else [])
    ))
    for p in pairs:
        ws.cell(r, 1, p).font = NORMAL_FONT
        ws.cell(r, 2, len(sw[sw['pair'] == p]) if not sw.empty else 0).font = NORMAL_FONT
        ws.cell(r, 3, len(dp[dp['pair'] == p]) if not dp.empty else 0).font = NORMAL_FONT
        ws.cell(r, 4, len(arb[arb['pair'] == p]) if not arb.empty else 0).font = NORMAL_FONT
        ws.cell(r, 5, len(sup[sup['pair'] == p]) if not sup.empty else 0).font = NORMAL_FONT
        for c in range(1, 6):
            ws.cell(r, c).border = THIN_BORDER
        r += 1

    _auto_width(ws)


# ── Sandwich Attacks ────────────────────────────────────────────

def _build_sandwich_sheet(wb, df):
    ws = wb.create_sheet('Sandwich Attacks')
    if df.empty:
        ws.cell(1, 1, 'No sandwich attacks detected.')
        return

    headers = [
        'ID', 'Pair', 'Block', 'Date/Time', 'Token0', 'Token1',
        'Attacker Address',
        'Front TX Hash', 'Front Sender (Orig Addr)', 'Front To (Dest Addr)',
        'Front Token0 In', 'Front Token1 In', 'Front Token0 Out', 'Front Token1 Out',
        'Front Gas (Gwei)',
        'Back TX Hash', 'Back Sender (Orig Addr)', 'Back To (Dest Addr)',
        'Back Token0 In', 'Back Token1 In', 'Back Token0 Out', 'Back Token1 Out',
        'Back Gas (Gwei)',
        'Victims', 'Gross Profit (USD)', 'Gas Cost (USD)', 'Net Profit (USD)',
    ]
    _write_header(ws, headers)

    for ri, (_, s) in enumerate(df.iterrows(), 2):
        cfg = PAIR_CONFIG.get(s['pair'], {})
        d0, d1 = cfg.get('token0_decimals', 18), cfg.get('token1_decimals', 18)
        vals = [
            s.get('sandwich_id', ri - 1), s['pair'], int(s['block_number']),
            _ts(s['timestamp']), cfg.get('token0_name', '?'), cfg.get('token1_name', '?'),
            s['attacker'],
            s['front_tx'], s.get('front_sender', ''), s.get('front_to', ''),
            _human(s['front_amount0_in'], d0), _human(s['front_amount1_in'], d1),
            _human(s['front_amount0_out'], d0), _human(s['front_amount1_out'], d1),
            round(s['front_gas_price'] / 1e9, 2),
            s['back_tx'], s.get('back_sender', ''), s.get('back_to', ''),
            _human(s['back_amount0_in'], d0), _human(s['back_amount1_in'], d1),
            _human(s['back_amount0_out'], d0), _human(s['back_amount1_out'], d1),
            round(s['back_gas_price'] / 1e9, 2),
            int(s['num_victims']),
            round(s.get('profit_usd', 0), 2),
            round(s.get('gas_cost_usd', 0), 2),
            round(s.get('net_profit_usd', 0), 2),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

        # colour net profit
        pc = ws.cell(ri, len(vals))
        pc.number_format = NUM_FMT_USD
        pc.font = PROFIT_POS if s.get('net_profit_usd', 0) >= 0 else PROFIT_NEG

        # hyperlinks
        ws.cell(ri, 8).hyperlink = _tx_url(s['front_tx'])
        ws.cell(ri, 8).font = LINK_FONT
        ws.cell(ri, 16).hyperlink = _tx_url(s['back_tx'])
        ws.cell(ri, 16).font = LINK_FONT

    _auto_width(ws)


# ── Victim Transactions ─────────────────────────────────────────

def _build_victim_sheet(wb, df):
    ws = wb.create_sheet('Victim Transactions')
    if df.empty:
        ws.cell(1, 1, 'No victim transactions.')
        return

    headers = [
        'Sandwich ID', 'Pair', 'Block', 'Date/Time',
        'Victim TX Hash', 'Victim Sender (Orig Addr)', 'Victim To (Dest Addr)',
        'Victim Entity', 'TX Index', 'Log Index', 'Direction',
        'Token0 In', 'Token1 In', 'Token0 Out', 'Token1 Out', 'Gas (Gwei)',
    ]
    _write_header(ws, headers)

    for ri, (_, v) in enumerate(df.iterrows(), 2):
        cfg = PAIR_CONFIG.get(v['pair'], {})
        d0, d1 = cfg.get('token0_decimals', 18), cfg.get('token1_decimals', 18)
        dir_lbl = 'Sell ' + cfg.get('token0_name', 'T0') if v['direction'] == 0 \
            else 'Sell ' + cfg.get('token1_name', 'T1')
        vals = [
            int(v['sandwich_id']), v['pair'], int(v['block_number']), _ts(v['timestamp']),
            v['victim_tx'], v.get('victim_sender', ''), v.get('victim_to', ''),
            v.get('victim_entity', ''),
            int(v['tx_index']), int(v['log_index']), dir_lbl,
            _human(v['amount0_in'], d0), _human(v['amount1_in'], d1),
            _human(v['amount0_out'], d0), _human(v['amount1_out'], d1),
            round(v['gas_price'] / 1e9, 2),
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(ri, ci, val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
        ws.cell(ri, 5).hyperlink = _tx_url(v['victim_tx'])
        ws.cell(ri, 5).font = LINK_FONT

    _auto_width(ws)


# ── Displacement ────────────────────────────────────────────────

def _build_displacement_sheet(wb, df):
    ws = wb.create_sheet('Displacement')
    if df.empty:
        ws.cell(1, 1, 'No displacement events detected.')
        return

    headers = [
        'Pair', 'Block', 'Date/Time', 'Direction',
        'Frontrunner TX', 'Frontrunner Sender (Orig)', 'Frontrunner To (Dest)',
        'Frontrunner Entity', 'Frontrunner Gas (Gwei)',
        'Victim TX', 'Victim Sender (Orig)', 'Victim To (Dest)',
        'Victim Entity', 'Victim Gas (Gwei)',
        'Gas Ratio', 'TX Gap',
        'Frontrunner Value (USD)', 'Victim Value (USD)',
    ]
    _write_header(ws, headers)

    for ri, (_, r) in enumerate(df.iterrows(), 2):
        cfg = PAIR_CONFIG.get(r['pair'], {})
        dir_lbl = 'Sell ' + cfg.get('token0_name', 'T0') if r['direction'] == 0 \
            else 'Sell ' + cfg.get('token1_name', 'T1')
        vals = [
            r['pair'], int(r['block_number']), _ts(r['timestamp']), dir_lbl,
            r['frontrunner_tx'], r.get('frontrunner_sender', ''), r.get('frontrunner_to', ''),
            r['frontrunner_entity'], round(r['frontrunner_gas_price'] / 1e9, 2),
            r['victim_tx'], r.get('victim_sender', ''), r.get('victim_to', ''),
            r['victim_entity'], round(r['victim_gas_price'] / 1e9, 2),
            r['gas_ratio'], int(r['tx_index_gap']),
            round(r.get('frontrunner_value_usd', 0), 2),
            round(r.get('victim_value_usd', 0), 2),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
        ws.cell(ri, 5).hyperlink = _tx_url(r['frontrunner_tx'])
        ws.cell(ri, 5).font = LINK_FONT
        ws.cell(ri, 10).hyperlink = _tx_url(r['victim_tx'])
        ws.cell(ri, 10).font = LINK_FONT

    _auto_width(ws)


# ── Arbitrage / Back-run ────────────────────────────────────────

def _build_arbitrage_sheet(wb, df):
    ws = wb.create_sheet('Arbitrage Back-run')
    if df.empty:
        ws.cell(1, 1, 'No arbitrage / back-run events detected.')
        return

    headers = [
        'Pair', 'Block', 'Date/Time',
        'Trigger TX', 'Trigger Sender (Orig)', 'Trigger To (Dest)',
        'Trigger Entity', 'Trigger Direction', 'Trigger Value (USD)',
        'Trigger Gas (Gwei)',
        'Back-runner TX', 'Back-runner Sender (Orig)', 'Back-runner To (Dest)',
        'Back-runner Entity', 'Back-runner Gas (Gwei)',
        'TX Gap', 'Net Profit (USD)', 'Gas Cost (USD)',
    ]
    _write_header(ws, headers)

    for ri, (_, r) in enumerate(df.iterrows(), 2):
        cfg = PAIR_CONFIG.get(r['pair'], {})
        dir_lbl = 'Sell ' + cfg.get('token0_name', 'T0') if r['trigger_direction'] == 0 \
            else 'Sell ' + cfg.get('token1_name', 'T1')
        vals = [
            r['pair'], int(r['block_number']), _ts(r['timestamp']),
            r['trigger_tx'], r.get('trigger_sender', ''), r.get('trigger_to', ''),
            r['trigger_entity'], dir_lbl,
            round(r.get('trigger_value_usd', 0), 2),
            round(r['trigger_gas_price'] / 1e9, 2),
            r['backrunner_tx'], r.get('backrunner_sender', ''), r.get('backrunner_to', ''),
            r['backrunner_entity'], round(r['backrunner_gas_price'] / 1e9, 2),
            int(r['tx_index_gap']),
            round(r.get('net_profit_usd', 0), 2),
            round(r.get('gas_cost_usd', 0), 2),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

        pc = ws.cell(ri, 17)
        pc.number_format = NUM_FMT_USD
        pc.font = PROFIT_POS if r.get('net_profit_usd', 0) >= 0 else PROFIT_NEG

        ws.cell(ri, 4).hyperlink = _tx_url(r['trigger_tx'])
        ws.cell(ri, 4).font = LINK_FONT
        ws.cell(ri, 11).hyperlink = _tx_url(r['backrunner_tx'])
        ws.cell(ri, 11).font = LINK_FONT

    _auto_width(ws)


# ── Suppression ─────────────────────────────────────────────────

def _build_suppression_sheet(wb, df):
    ws = wb.create_sheet('Suppression')
    if df.empty:
        ws.cell(1, 1, 'No suppression events detected.')
        return

    headers = [
        'Pair', 'Block', 'Date/Time',
        'Suppressor Entity', 'Suppressor TX Count',
        'Suppressor Gas Median (Gwei)', 'Block Gas Median (Gwei)',
        'Gas Premium (×)', 'Other Entities in Block', 'Others Gas Median (Gwei)',
    ]
    _write_header(ws, headers)

    for ri, (_, r) in enumerate(df.iterrows(), 2):
        vals = [
            r['pair'], int(r['block_number']), _ts(r['timestamp']),
            r['suppressor_entity'], int(r['suppressor_tx_count']),
            r['suppressor_gas_median_gwei'], r['block_gas_median_gwei'],
            r['gas_premium'], int(r['other_entities_in_block']),
            r['others_gas_median_gwei'],
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

    _auto_width(ws)
